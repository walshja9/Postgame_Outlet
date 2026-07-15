#!/usr/bin/env python3
"""
NFL Power Ratings builder (2026-2027 season).

Reads data/ratings.csv + data/config.csv, computes each team's rating as the
sum of its unit components, and writes a formatted Excel workbook with:
  - TEAM RATINGS   : every team sorted best-to-worst, components + total
  - QBs            : QB values alone (the dominant lever), sorted
  - Projections    : pick any two teams -> projected margin incl. home field
  - Inputs         : a read-only echo of the raw data for reference

The model (all values are points relative to a league-average team = 0.0):

    team_rating = qb_value + off_value + def_value
                  + coaching_value + scheme_value + your_edge

Edit data/ratings.csv to change the numbers, then re-run this script.
"""

import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "data")

# Component columns that sum into the team rating.
COMPONENTS = ["qb_value", "off_value", "def_value",
              "coaching_value", "scheme_value", "your_edge"]


def load_config():
    cfg = {}
    with open(os.path.join(DATA, "config.csv"), newline="") as f:
        for row in csv.DictReader(f):
            cfg[row["key"]] = row["value"]
    return cfg


def load_qb_depth():
    """Backup QBs (2nd/3rd string). Optional file; returns [] if absent."""
    path = os.path.join(DATA, "qb_depth.csv")
    if not os.path.exists(path):
        return []
    rows = []
    with open(path, newline="") as f:
        for r in csv.DictReader(f):
            try:
                r["value"] = float(r["value"] or 0)
            except ValueError:
                r["value"] = 0.0
            try:
                r["string"] = int(r["string"])
            except (ValueError, KeyError):
                r["string"] = 0
            rows.append(r)
    return rows


def load_teams():
    with open(os.path.join(DATA, "ratings.csv"), newline="") as f:
        rows = list(csv.DictReader(f))
    for r in rows:
        for c in COMPONENTS:
            try:
                r[c] = float(r[c] or 0)
            except ValueError:
                r[c] = 0.0
        r["rating"] = round(sum(r[c] for c in COMPONENTS), 2)
    return rows


# ---- styling helpers --------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")  # pale yellow = needs review
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center")


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_team_ratings(wb, teams):
    ws = wb.create_sheet("TEAM RATINGS")
    headers = ["Rank", "Team", "Conf", "Div", "QB", "QB val",
               "Offense", "Defense", "Coaching", "Scheme", "Edge",
               "RATING", "Review?"]
    ws.append(headers)
    style_header(ws, len(headers))
    for rank, t in enumerate(sorted(teams, key=lambda x: -x["rating"]), 1):
        ws.append([rank, t["team"], t["conf"], t["division"],
                   t["qb_name"], t["qb_value"], t["off_value"], t["def_value"],
                   t["coaching_value"], t["scheme_value"], t["your_edge"],
                   t["rating"], t["needs_review"]])
        if (t.get("needs_review") or "").strip().upper() == "Y":
            for c in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=c).fill = REVIEW_FILL
    # bold the RATING column
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=12).font = Font(bold=True)
    ws.freeze_panes = "A2"
    autosize(ws, [5, 24, 6, 7, 20, 8, 9, 9, 10, 9, 7, 9, 9])
    return ws


def build_qbs(wb, teams):
    ws = wb.create_sheet("QBs")
    headers = ["Rank", "QB", "Team", "QB value (pts vs avg)", "Review?"]
    ws.append(headers)
    style_header(ws, len(headers))
    for rank, t in enumerate(sorted(teams, key=lambda x: -x["qb_value"]), 1):
        ws.append([rank, t["qb_name"], t["team"], t["qb_value"], t["needs_review"]])
    ws.freeze_panes = "A2"
    autosize(ws, [5, 22, 24, 22, 9])
    return ws


def build_qb_depth(wb, depth):
    ws = wb.create_sheet("QB Depth")
    headers = ["Rank", "QB", "Team", "String", "Value", "Notes"]
    ws.append(headers)
    style_header(ws, len(headers))
    ordered = sorted(depth, key=lambda x: (-x["value"], x["string"]))
    for rank, d in enumerate(ordered, 1):
        ws.append([rank, d["qb_name"], d["team"],
                   f'{d["string"]}nd' if d["string"] == 2 else f'{d["string"]}rd',
                   d["value"], d.get("notes", "")])
    ws.freeze_panes = "A2"
    autosize(ws, [5, 22, 24, 7, 8, 34])
    return ws


def build_projections(wb, teams, hfa):
    """A live what-if sheet: choose Home/Away team, get projected margin.

    Uses Excel formulas + VLOOKUP into TEAM RATINGS so the user can change the
    two dropdowns without re-running Python.
    """
    ws = wb.create_sheet("Projections")
    n = len(teams)
    # TEAM RATINGS layout: Team in col B (2), RATING in col L (12).
    lookup_range = f"'TEAM RATINGS'!$B$2:$L${n + 1}"
    names = sorted(t["team"] for t in teams)

    ws["A1"] = "Game Projection"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "Home team"
    ws["A4"] = "Away team"
    ws["A6"] = "Home field adj"
    ws["A8"] = "Projected margin"
    ws["A9"] = "Result"
    for r in (3, 4, 6, 8, 9):
        ws[f"A{r}"].font = Font(bold=True)

    ws["B3"] = names[0]
    ws["B4"] = names[1]
    ws["B6"] = float(hfa)

    # ratings via vlookup (RATING is the 11th column counting from B)
    ws["C3"] = "=VLOOKUP(B3," + lookup_range + ",11,FALSE)"
    ws["C4"] = "=VLOOKUP(B4," + lookup_range + ",11,FALSE)"
    ws["C3"].number_format = "0.0"
    ws["C4"].number_format = "0.0"
    ws["B3"].comment = None

    # margin = home_rating - away_rating + hfa  (positive => home favored)
    ws["B8"] = "=C3-C4+B6"
    ws["B8"].number_format = "+0.0;-0.0"
    ws["B8"].font = Font(bold=True, size=12)
    ws["B9"] = ('=IF(B8>0,B3&" by "&TEXT(ABS(B8),"0.0"),'
                'B4&" by "&TEXT(ABS(B8),"0.0"))')

    # dropdowns
    dv = DataValidation(type="list",
                        formula1='"' + ",".join(names) + '"',
                        allow_blank=False)
    # Excel list validations have a ~255 char limit; if names exceed it,
    # fall back to a range-based list on a helper area.
    if len(",".join(names)) > 250:
        col = 8  # H
        for i, nm in enumerate(names, start=1):
            ws.cell(row=i, column=col, value=nm)
        rng = f"$H$1:$H${len(names)}"
        dv = DataValidation(type="list", formula1=rng, allow_blank=False)
        ws.column_dimensions["H"].hidden = True
    ws.add_data_validation(dv)
    dv.add(ws["B3"])
    dv.add(ws["B4"])

    autosize(ws, [16, 22, 10, 4, 4, 4, 4, 20])
    ws["A11"] = "Tip: change the Home/Away dropdowns; margin recalculates."
    ws["A11"].font = Font(italic=True, color="808080")
    return ws


def build_inputs(wb, teams):
    ws = wb.create_sheet("Inputs (raw)")
    headers = ["Team", "Conf", "Div", "QB", "QB val", "Off", "Def",
               "HC", "Coaching", "OC", "DC", "Scheme", "Edge", "Review?", "Notes"]
    ws.append(headers)
    style_header(ws, len(headers))
    for t in teams:
        ws.append([t["team"], t["conf"], t["division"], t["qb_name"],
                   t["qb_value"], t["off_value"], t["def_value"],
                   t["hc_name"], t["coaching_value"], t["oc_name"],
                   t["dc_name"], t["scheme_value"], t["your_edge"],
                   t["needs_review"], t.get("notes", "")])
    ws.freeze_panes = "A2"
    autosize(ws, [24, 6, 7, 18, 8, 7, 7, 18, 10, 18, 18, 9, 7, 9, 30])
    return ws


def main():
    cfg = load_config()
    teams = load_teams()
    depth = load_qb_depth()
    hfa = float(cfg.get("home_field_adv", 2.0))
    season = cfg.get("season", "2026")

    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet
    build_team_ratings(wb, teams)
    build_qbs(wb, teams)
    if depth:
        build_qb_depth(wb, depth)
    build_projections(wb, teams, hfa)
    build_inputs(wb, teams)

    out = os.path.join(HERE, f"NFL_Power_Ratings_{season}.xlsx")
    wb.save(out)
    reviewed = sum(1 for t in teams
                   if (t.get("needs_review") or "").strip().upper() != "Y")
    print(f"Wrote {out}")
    print(f"  {len(teams)} teams | {reviewed} confirmed, "
          f"{len(teams) - reviewed} still need review")
    print(f"  Home-field adj: {hfa}")


if __name__ == "__main__":
    main()
